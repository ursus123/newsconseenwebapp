"""
Greenfield Primary School — Newsconseen Demo Dataset
Diagnostic workbook: surfaces hidden anomalies for agents, copilot, and all user roles.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random

WB = openpyxl.Workbook()
WB.remove(WB.active)

# ── Colours ────────────────────────────────────────────────────────────────
H_FILL  = PatternFill("solid", fgColor="1E293B")   # slate-800
H2_FILL = PatternFill("solid", fgColor="0F766E")   # teal-700
WARN    = PatternFill("solid", fgColor="FEF9C3")   # yellow-100  (hidden story rows)
ERR     = PatternFill("solid", fgColor="FEE2E2")   # red-100
OK      = PatternFill("solid", fgColor="DCFCE7")   # green-100
H_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD    = Font(bold=True, size=9)
NORM    = Font(size=9)
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_sheet(name, headers, rows, notes=None, highlight_rows=None):
    ws = WB.create_sheet(name)
    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CENTER; cell.border = BORDER
    # Data rows
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = NORM; cell.alignment = LEFT; cell.border = BORDER
        if highlight_rows and (r - 2) in highlight_rows:
            fill = highlight_rows[r - 2]
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill
    # Auto-width
    for c in range(1, len(headers) + 1):
        col_letter = get_column_letter(c)
        max_len = len(str(headers[c - 1]))
        for r2 in range(2, min(len(rows) + 2, 50)):
            v = ws.cell(row=r2, column=c).value
            if v:
                max_len = max(max_len, min(len(str(v)), 40))
        ws.column_dimensions[col_letter].width = max_len + 2
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    if notes:
        ws.cell(row=len(rows) + 3, column=1, value="NOTES").font = BOLD
        ws.cell(row=len(rows) + 4, column=1, value=notes).font = Font(italic=True, size=8, color="64748B")
    return ws

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 0 — README
# ═══════════════════════════════════════════════════════════════════════════
ws_r = WB.create_sheet("README", 0)
ws_r.column_dimensions["A"].width = 18
ws_r.column_dimensions["B"].width = 80
readme = [
    ("School", "Greenfield Primary School"),
    ("Location", "Lusaka, Zambia"),
    ("Grades", "1 – 7  (approx 11 students per grade = 80 students)"),
    ("Currency", "ZMW (Zambian Kwacha)"),
    ("Purpose", "Newsconseen diagnostic dataset — every hidden story tests a different agent/copilot/UI path"),
    ("", ""),
    ("HIDDEN STORIES", ""),
    ("HS-1 Fleet overdue", "Bus 3 (2009 Toyota) service overdue 3 months — InventoryAgent + ComplianceAgent"),
    ("HS-2 Partial fees",  "Amara Dube (student) partial payer 3 consecutive terms — ComplianceAgent"),
    ("HS-3 Double salary", "Joseph Ndlovu (teacher) paid twice in Feb 2026 — AnomalyAgent"),
    ("HS-4 Low stock",     "Grade 7 Maths textbook qty=3 reorder=20 — InventoryAgent"),
    ("HS-5 Expired chem",  "Ethanol 95% in lab: expiry_date 2025-12-01 — ComplianceAgent"),
    ("HS-6 Paid not recv", "EduSupply Zambia: 3 purchase orders paid, goods receipt tasks incomplete — AnomalyAgent"),
    ("HS-7 Skills gap",    "Mrs. Banda (substitute) covers field trip without first_aid cert — ComplianceAgent"),
    ("HS-8 Overload",      "Mr. Tembo: 18 open tasks vs team avg 4 — OperationsAgent"),
    ("HS-9 Absenteeism",   "Mrs. Chirwa: 6 leave tasks in 90 days — AlertsEngine"),
    ("HS-10 Outcome drop", "Grade 5 task completion drops to 45% during Mrs. Chirwa leave periods — Copilot"),
    ("HS-11 Dep violation","Exam invigilation Gr7 = in_progress but Exam paper printing still incomplete — ComplianceAgent"),
    ("HS-12 Double fee",   "Chanda Mwale paid Term 2 tuition twice — AnomalyAgent"),
    ("", ""),
    ("SHEETS", ""),
    ("Staff",                   "30 staff — teachers, admin, support, drivers"),
    ("Students",                "80 students across Grades 1–7"),
    ("Contacts",                "15 external — suppliers, donors, gov, parent reps"),
    ("Enterprises",             "8 — school HQ, grade blocks, depts, 2 suppliers"),
    ("Products_Stock",          "40 — textbooks, stationery, lab, cleaning, sports"),
    ("Products_Fleet_Assets",   "16 — 5 buses, van, computers, projectors, equipment"),
    ("Tasks",                   "120 — teaching, admin, maintenance, events, leave, shifts"),
    ("Transactions_Fees",       "240 — student fee transactions (3 terms, 3 types, subset of 80)"),
    ("Transactions_Payroll",    "90 — staff salary Jan-Mar 2026"),
    ("Transactions_Procurement","45 — supplier purchase orders"),
    ("Relationships",           "50 — task deps, staff-class, student-guardian, supervision"),
    ("Addresses",               "14 — school buildings + supplier + staff addresses"),
]
ws_r.cell(1, 1, "FIELD").fill = H_FILL; ws_r.cell(1,1).font = H_FONT
ws_r.cell(1, 2, "VALUE").fill = H_FILL; ws_r.cell(1,2).font = H_FONT
for i, (k, v) in enumerate(readme, 2):
    ws_r.cell(i, 1, k).font = BOLD if k else NORM
    ws_r.cell(i, 2, v).font = NORM
ws_r.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — STAFF  (30 people, person_type=staff)
# ═══════════════════════════════════════════════════════════════════════════
STAFF_HEADERS = [
    "first_name","last_name","person_type","person_subtype","primary_role",
    "engagement_model","status","availability_status","gender",
    "date_of_birth","phone","email","start_date","cost_rate","payment_type",
    "skills","certification_name","certification_expiry","weekly_hours",
    "organization","internal_notes"
]
# (first, last, subtype, role, model, status, avail, gender, dob, phone, email, start, rate, pay_type, skills, cert, cert_exp, hrs, notes)
staff_raw = [
    # Principal / Admin
    ("Margaret","Banda","Principal","School Principal","employed","active","available","female","1972-03-14","+260-97-1234001","m.banda@greenfieldprimary.zm","2015-01-05",15000,"monthly_salary","leadership,administration,first_aid","First Aid Level 2","2027-06-01",40,"HS-9 Note: her deputy covers Grade5"),
    ("Charles","Mwansa","Vice Principal","Deputy Principal","employed","active","available","male","1975-07-22","+260-97-1234002","c.mwansa@greenfieldprimary.zm","2016-03-01",12000,"monthly_salary","leadership,curriculum,coaching","","",40,""),
    # Class teachers G1-G7
    ("Grace","Phiri","Class Teacher","Grade 1 Teacher","employed","active","available","female","1985-04-11","+260-97-1234003","g.phiri@greenfieldprimary.zm","2018-02-01",9000,"monthly_salary","first_aid,literacy_support,child_development","First Aid Level 1","2026-09-01",40,"Grade 1 class teacher"),
    ("Joseph","Ndlovu","Class Teacher","Grade 2 Teacher","employed","active","available","male","1983-09-05","+260-97-1234004","j.ndlovu@greenfieldprimary.zm","2017-05-01",9000,"monthly_salary","mathematics,science","","",40,"HS-3: Paid twice Feb 2026"),
    ("Esther","Tembo","Class Teacher","Grade 3 Teacher","employed","active","available","female","1990-01-30","+260-97-1234005","e.tembo@greenfieldprimary.zm","2019-09-01",9000,"monthly_salary","arts,music,reading","","",40,""),
    ("Daniel","Zulu","Class Teacher","Grade 4 Teacher","employed","active","available","male","1988-11-17","+260-97-1234006","d.zulu@greenfieldprimary.zm","2018-01-15",9000,"monthly_salary","social_studies,english","","",40,""),
    ("Patricia","Chirwa","Class Teacher","Grade 5 Teacher","employed","on_leave","on_leave","female","1986-06-02","+260-97-1234007","p.chirwa@greenfieldprimary.zm","2016-08-01",9000,"monthly_salary","english,drama","","",40,"HS-9+HS-10: Chronic absenteeism, Grade 5 outcomes drop"),
    ("Simon","Tembo","Class Teacher","Grade 6 Teacher","employed","active","busy","male","1984-03-25","+260-97-1234008","s.tembo@greenfieldprimary.zm","2015-11-01",9000,"monthly_salary","mathematics,ict,project_management","","",40,"HS-8: 18 open tasks vs avg 4"),
    ("Loveness","Mulenga","Class Teacher","Grade 7 Teacher","employed","active","available","female","1981-08-19","+260-97-1234009","l.mulenga@greenfieldprimary.zm","2013-06-01",9000,"monthly_salary","mathematics,science,exam_coordination","","",40,""),
    # Subject teachers
    ("Francis","Kabwe","Subject Teacher","Science Teacher","employed","active","available","male","1987-12-01","+260-97-1234010","f.kabwe@greenfieldprimary.zm","2020-01-10",8500,"monthly_salary","chemistry,biology,laboratory_safety,first_aid","First Aid Level 1","2025-11-30",40,"HS-7: Assigned to field trip (has first_aid)"),
    ("Agnes","Mwale","Subject Teacher","English Teacher","employed","active","available","female","1989-05-14","+260-97-1234011","a.mwale@greenfieldprimary.zm","2019-04-01",8500,"monthly_salary","english,creative_writing,drama","","",40,""),
    ("Kelvin","Lungu","Subject Teacher","Physical Education","employed","active","available","male","1991-10-08","+260-97-1234012","k.lungu@greenfieldprimary.zm","2021-03-01",8500,"monthly_salary","sports,first_aid,coaching","First Aid Level 1","2026-12-01",40,""),
    # Support staff
    ("Beatrice","Sakala","Bursar","School Bursar","employed","active","available","female","1978-02-27","+260-97-1234013","b.sakala@greenfieldprimary.zm","2014-07-01",10000,"monthly_salary","accounting,finance,fee_collection","","",40,""),
    ("Peter","Chanda","Clerk","Administrative Clerk","employed","active","available","male","1992-04-16","+260-97-1234014","p.chanda@greenfieldprimary.zm","2020-09-01",7000,"monthly_salary","data_entry,records_management","","",40,""),
    ("Naomi","Mwamba","Librarian","School Librarian","employed","active","available","female","1983-07-03","+260-97-1234015","n.mwamba@greenfieldprimary.zm","2017-02-01",7500,"monthly_salary","library_management,cataloguing,reading_support","","",40,""),
    ("Moses","Bwalya","IT Officer","ICT Officer","employed","active","available","male","1993-11-22","+260-97-1234016","m.bwalya@greenfieldprimary.zm","2022-01-05",8000,"monthly_salary","it_support,networking,hardware_repair","CompTIA A+","2028-01-01",40,""),
    ("Rose","Nkonde","School Nurse","Health Officer","employed","active","available","female","1980-09-14","+260-97-1234017","r.nkonde@greenfieldprimary.zm","2016-05-01",8500,"monthly_salary","first_aid,nursing,medication_administration,health_screening","Registered Nurse","2027-03-01",40,""),
    ("James","Mpundu","Security Guard","Security Officer","employed","active","available","male","1976-01-11","+260-97-1234018","j.mpundu@greenfieldprimary.zm","2013-09-01",4000,"monthly_salary","security,surveillance","","",48,""),
    ("Alice","Musonda","Cleaner","Cleaning Staff","employed","active","available","female","1982-05-30","+260-97-1234019","","2015-03-01",4500,"monthly_salary","cleaning,sanitation","","",40,""),
    ("Henry","Kafwanka","Cleaner","Cleaning Staff","employed","active","available","male","1979-08-06","+260-97-1234020","","2016-01-01",4500,"monthly_salary","cleaning,sanitation","","",40,""),
    # Bus drivers
    ("Edwin","Mwila","Bus Driver","Senior Driver","employed","active","available","male","1977-04-19","+260-97-1234021","e.mwila@greenfieldprimary.zm","2012-06-01",6000,"monthly_salary","driving,bus_driving,vehicle_maintenance,first_aid","Professional Driving Permit","2026-06-01",45,"HS-2: Drives Bus 3 (overdue service)"),
    ("Godfrey","Siame","Bus Driver","Driver","employed","active","available","male","1985-02-23","+260-97-1234022","","2018-09-01",6000,"monthly_salary","driving,bus_driving","Professional Driving Permit","2025-09-01",45,"Note: PDP expired — flag for renewal"),
    # Substitute / part-time
    ("Miriam","Banda","Relief Teacher","Supply Teacher","contracted","active","available","female","1994-07-12","+260-97-1234023","m.bandarel@greenfieldprimary.zm","2023-01-01",4500,"per_task","english,social_studies","","",0,"HS-7: Covers Grade5 during Chirwa leave — no first_aid cert"),
    ("Tendai","Phiri","Relief Teacher","Supply Teacher","contracted","active","available","male","1996-03-04","+260-97-1234024","","2024-02-01",4500,"per_task","mathematics,science","","",0,""),
    # Kitchen
    ("Florence","Mutale","Cook","Kitchen Staff","employed","active","available","female","1975-10-17","+260-97-1234025","","2014-01-01",4000,"monthly_salary","food_preparation,hygiene,nutrition","Food Hygiene Cert","2026-08-01",40,""),
    ("Bertha","Ngoma","Cook","Kitchen Staff","employed","active","available","female","1980-06-28","+260-97-1234026","","2017-03-01",4000,"monthly_salary","food_preparation,hygiene","","",40,""),
    # Grounds
    ("Elias","Mumba","Groundsman","Groundskeeper","employed","active","available","male","1971-12-01","+260-97-1234027","","2010-08-01",3500,"monthly_salary","grounds_maintenance,plumbing,general_repairs","","",40,""),
    # Laboratory tech
    ("Susan","Kaputo","Lab Technician","Laboratory Technician","employed","active","available","female","1988-09-09","+260-97-1234028","s.kaputo@greenfieldprimary.zm","2021-06-01",7500,"monthly_salary","laboratory_safety,chemistry,biology,equipment_maintenance","Lab Safety Level 2","2026-04-01",40,"HS-5: Manages expired Ethanol — should have flagged it"),
    # Tuck shop
    ("Charity","Mbewe","Tuck Shop Attendant","Canteen Operator","employed","active","available","female","1983-03-21","+260-97-1234029","","2019-01-01",3500,"monthly_salary","cash_handling,stock_management","","",40,""),
    # Exam coordinator (also Grade7 teacher — dual role)
    ("Victor","Chisanga","Exam Coordinator","Examination Officer","employed","active","available","male","1979-11-11","+260-97-1234030","v.chisanga@greenfieldprimary.zm","2011-04-01",9500,"monthly_salary","exam_coordination,data_analysis,administration","","",40,""),
]
staff_rows = [
    (r[0],r[1],"staff",r[2],r[3],r[4],r[5],r[6],r[7],r[8],r[9],r[10],r[11],r[12],r[13],r[14],r[15],r[16],r[17],"Greenfield Primary School",r[18])
    for r in staff_raw
]
# Highlight hidden-story rows
staff_highlights = {
    3: WARN,   # Joseph Ndlovu double salary
    6: ERR,    # Patricia Chirwa absenteeism / on_leave
    7: WARN,   # Simon Tembo overloaded
    20: WARN,  # Edwin Mwila drives overdue bus
    21: ERR,   # Godfrey Siame expired PDP
    22: WARN,  # Miriam Banda — skills gap (no first_aid)
}
make_sheet("Staff", STAFF_HEADERS, staff_rows, highlight_rows=staff_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — STUDENTS (80 students, person_type=client)
# ═══════════════════════════════════════════════════════════════════════════
STU_HEADERS = [
    "first_name","last_name","person_type","person_subtype","primary_role",
    "engagement_model","status","gender","date_of_birth",
    "phone","email","start_date","organization","internal_notes"
]

def student_dob(grade):
    base = 2026 - (grade + 6)
    return f"{base}-{random.randint(1,12):02d}-{random.randint(1,28):02d}"

GRADE_CLASSES = {
    1: ["Aisha","Bwalya","Chama","Daliso","Esme","Francis","Grace","Hector","Irene","Jerome","Kunda"],
    2: ["Lena","Mable","Nathan","Olivia","Paul","Queen","Ronald","Stella","Tanya","Unity","Victor"],
    3: ["Wanda","Xavier","Yvonne","Zain","Abel","Brenda","Caesar","Diana","Edgar","Faith","George"],
    4: ["Hannah","Isaac","Julia","Kenneth","Laura","Martin","Nancy","Oscar","Priscilla","Quinn","Roy"],
    5: ["Sally","Thomas","Ursula","Vincent","Wendy","Alex","Bongi","Chrissa","Dumisani","Elina","Fred"],
    6: ["Gloria","Hosea","Ivy","James","Karen","Lucas","Miriam","Noah","Ophelia","Patrick","Rachel"],
    7: ["Sarah","Timothy","Ulrike","Vasco","Winnie","Xola","Yolanda","Zola","Aaron","Bertha","Chanda"],
}
LAST_NAMES = ["Phiri","Banda","Mwale","Tembo","Zulu","Mwamba","Lungu","Kabwe","Nkonde","Mwila",
              "Mulenga","Sakala","Chanda","Bwalya","Musonda","Mumba","Mutale","Kaputo","Mbewe","Siame",
              "Chirwa","Ndlovu","Mpundu","Kafwanka","Ngoma","Chisanga","Mwansa","Kaunda","Dube","Ncube"]

students_rows = []
student_highlights = {}
stu_idx = 0
for grade, names in GRADE_CLASSES.items():
    last_names_used = random.sample(LAST_NAMES, len(names))
    for i, first in enumerate(names):
        last = last_names_used[i]
        dob = student_dob(grade)
        gender = "female" if first in ["Aisha","Esme","Grace","Irene","Lena","Mable","Olivia","Stella","Tanya","Unity","Wanda","Yvonne","Diana","Faith","Hannah","Julia","Laura","Nancy","Priscilla","Sally","Ursula","Wendy","Brenda","Gloria","Ivy","Karen","Miriam","Ophelia","Rachel","Sarah","Ulrike","Winnie","Yolanda","Bertha","Elina","Chrissa","Bongi"] else "male"
        note = f"Grade {grade}"
        # Special students
        if first == "Aisha" and grade == 1:
            note += " | New intake Term 1 2026"
        if first == "Chanda" and grade == 7:
            note += " | HS-12: Double fee payment Term 2"
            student_highlights[stu_idx] = WARN
        if first == "Winnie" and grade == 7:
            note += " | HS-2: Amara Dube alias — partial fee payer 3 terms"
            first, last = "Amara", "Dube"
            student_highlights[stu_idx] = ERR
        students_rows.append((
            first, last, "client", f"Grade {grade} Student", f"Grade {grade}",
            "enrolled", "active", gender, dob,
            "", "", f"2024-01-15", "Greenfield Primary School", note
        ))
        stu_idx += 1

make_sheet("Students", STU_HEADERS, students_rows, highlight_rows=student_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — CONTACTS (15 external contacts)
# ═══════════════════════════════════════════════════════════════════════════
CONT_HEADERS = [
    "first_name","last_name","person_type","person_subtype","primary_role",
    "engagement_model","status","gender","phone","email","organization","internal_notes"
]
contacts = [
    ("David","Mutale","contact","Supplier","Sales Manager","contracted","active","male","+260-96-5551001","d.mutale@edusupply.zm","EduSupply Zambia Ltd","HS-6: 3 POs paid, goods not received"),
    ("Christine","Bwalya","contact","Supplier","Account Manager","contracted","active","female","+260-96-5551002","c.bwalya@officeplus.zm","Office Plus Zambia","Stationery supplier — reliable"),
    ("Robert","Phiri","contact","Supplier","Director","contracted","active","male","+260-96-5551003","r.phiri@techsupplies.zm","Tech Supplies Africa","Computer equipment supplier"),
    ("Jennifer","Kabwe","contact","Government Official","District Education Officer","appointed","active","female","+260-97-7771001","j.kabwe@moe.gov.zm","Ministry of Education Lusaka","Annual school inspection contact"),
    ("Samuel","Lungu","contact","Government Official","Health Inspector","appointed","active","male","+260-97-7771002","s.lungu@moh.gov.zm","Ministry of Health","Annual health inspection"),
    ("Mary","Banda","contact","Donor","Program Officer","contracted","active","female","+260-96-8881001","m.banda@unicef.org","UNICEF Zambia","School supplies donor"),
    ("Thomas","Mwale","contact","Donor","Country Director","appointed","active","male","+260-96-8881002","t.mwale@savechildren.org","Save the Children Zambia","Reading program sponsor"),
    ("Alice","Tembo","contact","Parent Representative","PTA Chairperson","volunteer","active","female","+260-97-9991001","","","PTA Chair — Term 1-2026"),
    ("George","Zulu","contact","Parent Representative","PTA Secretary","volunteer","active","male","+260-97-9991002","","","PTA Secretary"),
    ("Brenda","Mulenga","contact","Parent Representative","Class Rep Grade 7","volunteer","active","female","+260-97-9991003","","","Grade 7 Parent Rep"),
    ("Patrick","Chanda","contact","External Examiner","Examination Board Rep","appointed","active","male","+260-97-7772001","p.chanda@examboard.zm","Examinations Council Zambia","Grade 7 exams coordinator"),
    ("Nancy","Sakala","contact","Medical Supplier","Pharmacist","contracted","active","female","+260-96-5552001","n.sakala@medsupply.zm","MedSupply Zambia","First aid supplies vendor"),
    ("Victor","Mumba","contact","Transport Authority","Inspector","appointed","active","male","+260-97-7773001","v.mumba@rtsa.zm","Road Transport Safety Authority","Bus roadworthiness inspections"),
    ("Eunice","Kaputo","contact","Bank","Relationship Manager","contracted","active","female","+260-97-8882001","e.kaputo@zanaco.zm","ZANACO Bank Zambia","School banking contact"),
    ("Arnold","Mwamba","contact","Security Company","Operations Manager","contracted","active","male","+260-96-5553001","a.mwamba@guardsecurity.zm","Guard & Secure Ltd","Security services contractor"),
]
make_sheet("Contacts", CONT_HEADERS, contacts, highlight_rows={0: WARN})

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — ENTERPRISES
# ═══════════════════════════════════════════════════════════════════════════
ENT_HEADERS = [
    "enterprise_name","enterprise_type","enterprise_tier","enterprise_subtype",
    "parent_enterprise_id","status","operating_status","phone","email","website",
    "city","country","naics_code","naics_title","description","internal_notes"
]
enterprises = [
    ("Greenfield Primary School","nonprofit","headquarters","Primary School","","active","open","+260-211-123456","info@greenfieldprimary.zm","www.greenfieldprimary.zm","Lusaka","Zambia","611110","Elementary and Secondary Schools","Main school headquarters — Grades 1 to 7","Founded 2010; MoE License: ZM-EDU-2010-0042"),
    ("Grade 1-3 Block","nonprofit","department","Junior Section","Greenfield Primary School","active","open","","","","Lusaka","Zambia","611110","Elementary and Secondary Schools","Junior grades block — Grades 1, 2, 3","Capacity: 36 pupils"),
    ("Grade 4-6 Block","nonprofit","department","Middle Section","Greenfield Primary School","active","open","","","","Lusaka","Zambia","611110","Elementary and Secondary Schools","Middle grades block — Grades 4, 5, 6","Capacity: 36 pupils"),
    ("Grade 7 Block","nonprofit","department","Senior Section","Greenfield Primary School","active","open","","","","Lusaka","Zambia","611110","Elementary and Secondary Schools","Senior grade block — Grade 7 (exam class)","Capacity: 11 pupils; External exams venue"),
    ("Administration Block","nonprofit","department","Administration","Greenfield Primary School","active","open","","","","Lusaka","Zambia","611110","Elementary and Secondary Schools","Principal office, bursar, staffroom","Houses finance records and staff welfare"),
    ("Transport Department","nonprofit","unit","Fleet Management","Greenfield Primary School","active","open","","","","Lusaka","Zambia","485410","School and Employee Bus Transportation","Manages 5 buses and 1 van","HS-1: Bus 3 overdue service"),
    ("Science & Computer Lab","nonprofit","unit","Laboratory","Greenfield Primary School","active","open","","","","Lusaka","Zambia","611110","Elementary and Secondary Schools","Combined science lab and computer lab","HS-5: Expired chemicals in lab"),
    ("EduSupply Zambia Ltd","commercial","headquarters","Educational Supplies","","active","open","+260-96-5551001","orders@edusupply.zm","","Lusaka","Zambia","424120","Stationery and Office Supplies Merchant","Textbook and supplies distributor","HS-6: Unreliable delivery — 3 unpaid goods receipts"),
]
ent_highlights = {6: WARN, 7: ERR}
make_sheet("Enterprises", ENT_HEADERS, enterprises, highlight_rows=ent_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 — PRODUCTS: STOCK (40 items)
# ═══════════════════════════════════════════════════════════════════════════
PROD_HEADERS = [
    "product_name","item_type","item_class","item_subtype","unit_of_measure",
    "stock_quantity","reorder_level","unit_cost","expiry_date","internal_notes"
]
products_stock = [
    # Textbooks — Grade 1-7 (Maths, English, Science, Social Studies)
    ("Grade 1 Mathematics Textbook","physical","non_perishable","Textbook","piece",45,10,85,"",""),
    ("Grade 1 English Textbook","physical","non_perishable","Textbook","piece",43,10,85,"",""),
    ("Grade 2 Mathematics Textbook","physical","non_perishable","Textbook","piece",38,10,85,"",""),
    ("Grade 2 English Textbook","physical","non_perishable","Textbook","piece",40,10,85,"",""),
    ("Grade 3 Mathematics Textbook","physical","non_perishable","Textbook","piece",35,10,85,"",""),
    ("Grade 3 Social Studies Textbook","physical","non_perishable","Textbook","piece",33,10,75,"",""),
    ("Grade 4 Mathematics Textbook","physical","non_perishable","Textbook","piece",28,10,90,"",""),
    ("Grade 4 Science Textbook","physical","non_perishable","Textbook","piece",30,10,90,"",""),
    ("Grade 5 Mathematics Textbook","physical","non_perishable","Textbook","piece",25,10,90,"",""),
    ("Grade 5 English Textbook","physical","non_perishable","Textbook","piece",24,10,90,"",""),
    ("Grade 6 Mathematics Textbook","physical","non_perishable","Textbook","piece",22,10,95,"",""),
    ("Grade 6 Science Textbook","physical","non_perishable","Textbook","piece",20,10,95,"",""),
    ("Grade 7 Mathematics Textbook","physical","non_perishable","Textbook","piece",3,20,95,"","HS-4: CRITICAL LOW STOCK — only 3 left, reorder=20, 11 students in Grade 7"),
    ("Grade 7 English Textbook","physical","non_perishable","Textbook","piece",11,10,95,"",""),
    ("Grade 7 Science Textbook","physical","non_perishable","Textbook","piece",10,10,95,"",""),
    # Stationery
    ("Exercise Books (A4 lined)","physical","consumable","Stationery","piece",320,50,8,"",""),
    ("Pens (Blue Biro)","physical","consumable","Stationery","piece",180,50,2,"",""),
    ("Pencils (HB)","physical","consumable","Stationery","piece",200,60,1.5,"",""),
    ("Rulers (30cm)","physical","non_perishable","Stationery","piece",65,20,5,"",""),
    ("Whiteboard Markers (Black)","physical","consumable","Stationery","piece",24,10,12,"",""),
    ("Whiteboard Markers (Colour Set)","physical","consumable","Stationery","piece",8,5,18,"",""),
    ("Manila Paper (Ream)","physical","consumable","Stationery","piece",12,5,45,"",""),
    # Lab supplies
    ("Ethanol 95% (1L bottle)","physical","hazardous","Lab Chemical","liter",4,2,85,"2025-12-01","HS-5: EXPIRED — expiry was 2025-12-01, still in lab storage"),
    ("Litmus Paper (Red/Blue)","physical","consumable","Lab Supply","piece",200,50,0.5,"2027-06-01",""),
    ("Safety Goggles","physical","reusable","Lab Supply","piece",18,10,35,"",""),
    ("Lab Coats (Small)","physical","reusable","Lab Supply","piece",12,8,80,"",""),
    # Cleaning
    ("Disinfectant (5L)","physical","consumable","Cleaning Supply","liter",15,5,65,"2027-01-01",""),
    ("Hand Soap (500ml)","physical","consumable","Cleaning Supply","piece",30,10,18,"",""),
    ("Mop Heads","physical","consumable","Cleaning Supply","piece",6,4,45,"",""),
    # Sports
    ("Soccer Ball","physical","non_perishable","Sports Equipment","piece",5,2,180,"",""),
    ("Netball","physical","non_perishable","Sports Equipment","piece",4,2,160,"",""),
    ("Skipping Ropes","physical","non_perishable","Sports Equipment","piece",15,5,25,"",""),
    ("Athletics Baton Set","physical","non_perishable","Sports Equipment","piece",3,1,120,"",""),
    # First aid
    ("Bandages (box of 50)","physical","consumable","First Aid Supply","piece",8,4,35,"2028-01-01",""),
    ("Antiseptic Cream (50g)","physical","perishable","First Aid Supply","piece",10,5,28,"2026-10-01",""),
    ("Paracetamol Tablets 500mg","physical","controlled","Medication","piece",2,5,45,"2026-07-01","Low stock — controlled medication"),
    # Art supplies (ordered, not yet received)
    ("Poster Paint Set","physical","consumable","Art Supply","piece",0,10,55,"","HS-6 Note: PO raised with EduSupply, not yet received — stock shows 0"),
    ("Drawing Paper A3 (Ream)","physical","consumable","Art Supply","piece",2,8,38,"","PO raised — short supply"),
    # Library
    ("Library Reference Books (set)","physical","non_perishable","Library Resource","piece",45,0,0,"","Donated by Save the Children"),
    ("Story Books (Grade 1-3 set)","physical","non_perishable","Library Resource","piece",60,0,0,"",""),
]
prod_highlights = {12: ERR, 22: ERR, 36: WARN, 37: WARN}
make_sheet("Products_Stock", PROD_HEADERS, products_stock, highlight_rows=prod_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6 — PRODUCTS: FLEET & ASSETS
# ═══════════════════════════════════════════════════════════════════════════
fleet = [
    # Buses
    ("School Bus 1 — Reg BAA 4521","physical","serialized","School Bus","piece",1,0,650000,"","Year: 2019 | Make: Toyota Coaster | Capacity: 29 | Condition: Good | Last service: 2025-11-01 | Next service due: 2026-05-01 | Status: Roadworthy"),
    ("School Bus 2 — Reg BAB 7733","physical","serialized","School Bus","piece",1,0,720000,"","Year: 2020 | Make: Toyota Coaster | Capacity: 29 | Condition: Good | Last service: 2025-12-15 | Next service due: 2026-06-15 | Status: Roadworthy"),
    ("School Bus 3 — Reg AAM 1198","physical","serialized","School Bus","piece",1,0,180000,"","HS-1: Year: 2009 | Make: Toyota Coaster | Capacity: 29 | Condition: Poor | Last service: 2025-10-01 | Next service due: 2026-01-01 | OVERDUE BY 3 MONTHS | Requires inspection before next use"),
    ("School Bus 4 — Reg BAC 2267","physical","serialized","School Bus","piece",1,0,420000,"","Year: 2015 | Make: Toyota Coaster | Capacity: 29 | Condition: Fair | Last service: 2025-09-01 | Next service due: 2026-03-01 | Due soon"),
    ("School Bus 5 — Reg BAD 5512","physical","serialized","School Bus","piece",1,0,550000,"","Year: 2017 | Make: Toyota Coaster | Capacity: 29 | Condition: Good | Last service: 2026-01-15 | Next service due: 2026-07-15"),
    ("School Van — Reg BAE 0091","physical","serialized","School Van","piece",1,0,280000,"","Year: 2021 | Make: Toyota HiAce | Capacity: 12 | Condition: Good | Last service: 2026-02-01 | Next service due: 2026-08-01"),
    # Computer lab
    ("Desktop Computer (Dell OptiPlex)","physical","serialized","Computer","piece",18,0,8500,"","Lab: 25 total; 18 working; 7 out of service — see maintenance tasks"),
    ("Desktop Computer (Old HP)","physical","serialized","Computer","piece",7,0,2000,"","Out of service — earmarked for disposal"),
    # Projectors
    ("Projector — Epson EB-X49","physical","serialized","Projector","piece",2,0,12000,"","2 units functional"),
    ("Projector — Optoma S341","physical","serialized","Projector","piece",1,0,3000,"","Out of service — lamp needs replacement"),
    # Lab equipment
    ("Microscope (Binocular)","physical","serialized","Lab Equipment","piece",6,0,15000,"","Science lab — all 6 functional"),
    ("Digital Weighing Scale","physical","serialized","Lab Equipment","piece",2,0,2200,"","Lab use"),
    ("Printer — HP LaserJet","physical","serialized","Printer","piece",3,0,6500,"","Admin: 2 working; 1 paper jam issue"),
    ("Photocopier — Ricoh IM 2702","physical","serialized","Photocopier","piece",1,0,45000,"","Admin block — main copier"),
    # Generator
    ("Generator — Kipor 5kVA","physical","serialized","Generator","piece",1,0,25000,"","Backup power — last service 2025-08-01"),
    # Sports
    ("Goal Posts (football)","physical","non_perishable","Sports Equipment","piece",2,0,8000,"","Playing field"),
]
fleet_highlights = {2: ERR, 3: WARN, 9: WARN}
make_sheet("Products_Fleet_Assets", PROD_HEADERS, fleet, highlight_rows=fleet_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7 — TASKS (120 tasks)
# ═══════════════════════════════════════════════════════════════════════════
TASK_HEADERS = [
    "title","task_type","status","priority","assigned_to",
    "due_date","outcome","notes"
]
tasks = []
task_highlights = {}
t_idx = 0

def add_task(title, ttype, status, priority, assigned, due, outcome="", notes="", flag=None):
    global t_idx
    tasks.append((title, ttype, status, priority, assigned, due, outcome, notes))
    if flag:
        task_highlights[t_idx] = flag
    t_idx += 1

# ── Teaching tasks ─────────────────────────────────────────────────────────
for grade in range(1, 8):
    teacher = ["Grace Phiri","Joseph Ndlovu","Esther Tembo","Daniel Zulu","Patricia Chirwa","Simon Tembo","Loveness Mulenga"][grade-1]
    add_task(f"Grade {grade} Weekly Lesson Plan — Term 2 Week 1","lesson_plan","completed","normal",teacher,"2026-01-30","submitted","")
    add_task(f"Grade {grade} Marking — Term 2 Week 1 Tests","marking","completed","normal",teacher,"2026-02-06","completed","")

# Grade 5 tasks during Chirwa leave — lower completion
for i in range(1, 5):
    flag = ERR if i % 2 == 0 else WARN
    add_task(f"Grade 5 Week {i+1} — Cover (supply teacher)","cover_lesson","incomplete" if i > 2 else "completed","normal","Miriam Banda",f"2026-02-{10+i*5:02d}","","HS-10: Grade 5 outcomes dropping during Chirwa absence",flag)

# ── Administrative tasks ────────────────────────────────────────────────────
add_task("Term 2 Fee Collection Report","admin","completed","high","Beatrice Sakala","2026-02-15","submitted","")
add_task("Procurement Order — EduSupply (Textbooks)","procurement","completed","high","Peter Chanda","2026-01-20","","HS-6: PO submitted and paid — goods not yet received",WARN)
add_task("Procurement Order — EduSupply (Art Supplies)","procurement","completed","normal","Peter Chanda","2026-01-25","","HS-6: PO submitted and paid — goods not received",WARN)
add_task("Procurement Order — EduSupply (Lab Supplies)","procurement","completed","normal","Susan Kaputo","2026-02-01","","HS-6: Paid — Ethanol delivered but EXPIRED stock accepted",ERR)
add_task("Goods Receipt — EduSupply Textbooks","goods_receipt","incomplete","high","Peter Chanda","2026-02-01","","HS-6: PO paid 2026-01-20, goods receipt still incomplete",ERR)
add_task("Goods Receipt — EduSupply Art Supplies","goods_receipt","incomplete","normal","Peter Chanda","2026-02-05","","HS-6: Still incomplete",ERR)
add_task("Goods Receipt — EduSupply Lab Supplies","goods_receipt","completed","normal","Susan Kaputo","2026-02-10","completed","Received expired Ethanol — should have been rejected",WARN)
add_task("Monthly Staff Payroll Processing","payroll","completed","high","Beatrice Sakala","2026-01-31","submitted","")
add_task("Monthly Staff Payroll Processing","payroll","completed","high","Beatrice Sakala","2026-02-28","submitted","HS-3: Joseph Ndlovu paid twice — duplicate entry",WARN)
add_task("Monthly Staff Payroll Processing","payroll","completed","high","Beatrice Sakala","2026-03-31","submitted","")
add_task("Annual School License Renewal — MoE","compliance","completed","urgent","Margaret Banda","2026-01-15","approved","")
add_task("Health & Safety Inspection Prep","compliance","completed","high","Rose Nkonde","2026-02-20","passed","")

# ── Simon Tembo overload (HS-8) ─────────────────────────────────────────────
for i in range(1, 14):
    prio = "high" if i <= 4 else "normal"
    status = "completed" if i <= 4 else "open"
    add_task(f"Grade 6 Extended Task #{i} — ICT/Admin","admin","open" if i > 4 else "completed",prio,"Simon Tembo",f"2026-04-{min(i*2,30):02d}","",f"HS-8: Tembo overloaded — task {i} of 18",WARN if i > 4 else None)

# ── Exam sequence (HS-11) ────────────────────────────────────────────────────
add_task("Grade 7 Exam Schedule Confirmed","exam_admin","completed","urgent","Victor Chisanga","2026-03-01","confirmed","Dependency: Exam paper printing and invigilation depend on this",OK)
add_task("Grade 7 Exam Paper Printing","exam_admin","incomplete","urgent","Victor Chisanga","2026-03-15","","HS-11: Must complete before invigilation — currently INCOMPLETE",ERR)
add_task("Grade 7 Exam Invigilation — Session 1","exam_admin","in_progress","urgent","Loveness Mulenga","2026-03-20","","HS-11: DEPENDENCY VIOLATION — invigilation in_progress but paper printing incomplete",ERR)
add_task("Grade 7 Exam Invigilation — Session 2","exam_admin","open","urgent","Victor Chisanga","2026-03-22","","Depends on: paper printing complete",WARN)
add_task("Grade 7 Exam Results Compilation","exam_admin","open","high","Loveness Mulenga","2026-04-05","","Depends on: both invigilation sessions",WARN)

# ── Field trip (HS-7) ────────────────────────────────────────────────────────
add_task("Grade 6 Field Trip — Science Museum Lusaka","field_trip","open","normal","Francis Kabwe","2026-04-15","","HS-7: Lead teacher Francis Kabwe (has first_aid) — if absent, Miriam Banda covers",OK)
add_task("Grade 6 Field Trip Cover — Supply Teacher","cover_lesson","open","normal","Miriam Banda","2026-04-15","","HS-7: SKILLS GAP — Miriam Banda has no first_aid certification but may cover field trip",ERR)
add_task("Grade 6 Field Trip Risk Assessment","compliance","completed","high","Margaret Banda","2026-04-01","approved","")
add_task("Grade 6 Field Trip Parent Consent Collection","admin","completed","high","Daniel Zulu","2026-04-08","completed","All 11 Grade 6 consent forms returned")

# ── Maintenance tasks ─────────────────────────────────────────────────────────
add_task("Bus 3 (AAM 1198) — Full Service Overdue","vehicle_maintenance","open","urgent","Edwin Mwila","2026-01-01","","HS-1: Service due 2026-01-01 — now 3+ months overdue. Bus still in daily use",ERR)
add_task("Bus 4 (BAC 2267) — Scheduled Service","vehicle_maintenance","open","high","Edwin Mwila","2026-03-01","","Service due March 2026",WARN)
add_task("Projector Lamp Replacement — Optoma S341","maintenance","open","normal","Moses Bwalya","2026-03-15","","Projector out of service until lamp replaced",WARN)
add_task("Computer Lab Repair — 7 HP units","maintenance","in_progress","high","Moses Bwalya","2026-02-28","","7 old HP desktops earmarked for disposal or repair",WARN)
add_task("Lab Chemical Audit & Disposal","compliance","open","urgent","Susan Kaputo","2026-04-01","","HS-5: Audit of expired Ethanol and other lab chemicals overdue",ERR)
add_task("Generator Service — Kipor 5kVA","maintenance","open","normal","Elias Mumba","2026-04-15","","Last service Aug 2025 — 8 months ago",WARN)

# ── Leave requests (HS-9 — Chirwa) ───────────────────────────────────────────
for month, days in [("January","13-17"),("February","03-07"),("February","24-28"),("March","10-14"),("March","24-28"),("April","07-11")]:
    add_task(f"Leave — Patricia Chirwa ({month})","leave_request","completed","normal","Patricia Chirwa",f"2026-{['Jan','Feb','Feb','Mar','Mar','Apr'].index(month[:3])+1:02d}-{days.split('-')[1]}","approved",f"HS-9: Leave #{tasks.__len__()} — pattern of 6 leaves in 90 days",ERR)

# ── Clock-in / shift tasks ─────────────────────────────────────────────────
for day in range(1, 8):
    add_task(f"Morning Duty — Gate Register 2026-04-{day:02d}","shift","completed","normal","James Mpundu",f"2026-04-{day:02d}","completed","Security gate register completed")

# ── Health & nutrition ────────────────────────────────────────────────────────
add_task("Term 2 Health Screening — All Students","health_check","in_progress","high","Rose Nkonde","2026-02-28","","Grades 1-4 done; Grades 5-7 pending")
add_task("First Aid Kit Restock","procurement","open","normal","Rose Nkonde","2026-04-10","","Bandages and antiseptic low")
add_task("Staff CPR Refresher Training","training","open","normal","Rose Nkonde","2026-04-20","","Annual CPR refresher for all staff",WARN)

# ── PTA / community tasks ──────────────────────────────────────────────────
add_task("Term 2 Parent-Teacher Meeting","meeting","completed","normal","Margaret Banda","2026-03-05","completed","80% parent attendance")
add_task("Annual Sports Day Planning","event","in_progress","normal","Kelvin Lungu","2026-04-25","","Sports day scheduled 2026-05-10")
add_task("School Newsletter — Term 2 Edition","communication","completed","normal","Peter Chanda","2026-03-15","published","")
add_task("Donation Acknowledgement — UNICEF","admin","completed","normal","Margaret Banda","2026-02-10","sent","Thank you letters sent")

make_sheet("Tasks", TASK_HEADERS, tasks, highlight_rows=task_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8 — TRANSACTIONS: FEES (student fees)
# ═══════════════════════════════════════════════════════════════════════════
FEE_HEADERS = [
    "transaction_type","category","reference_number","amount","amount_paid",
    "payment_status","due_date","person_name","description","notes"
]
fee_rows = []
fee_highlights = {}

terms = [
    ("Term 1 2026","2026-01-31"),
    ("Term 2 2026","2026-04-30"),
    ("Term 3 2025","2025-09-30"),
]
fee_students_sample = [
    # (first, last, grade, pay_tuition, pay_transport, pay_meals, special)
    ("Aisha","Phiri",1,True,True,True,None),
    ("Bwalya","Banda",1,True,False,True,None),
    ("Chama","Mwale",1,True,True,False,None),
    ("Lena","Tembo",2,True,True,True,None),
    ("Nathan","Zulu",2,True,False,False,None),
    ("Wanda","Mwamba",3,True,True,True,None),
    ("Hannah","Lungu",4,True,False,True,None),
    ("Sally","Kabwe",5,True,True,True,None),
    ("Thomas","Nkonde",5,True,False,True,None),
    ("Gloria","Mwila",6,True,True,True,None),
    ("James","Mulenga",6,True,False,False,None),
    ("Sarah","Sakala",7,True,True,True,None),
    ("Timothy","Chanda",7,True,True,True,None),
    ("Amara","Dube",7,True,True,False,"partial"),  # HS-2
    ("Chanda","Mwale",7,True,False,True,"double"),  # HS-12
    ("Winnie","Chirwa",7,True,True,True,None),
    ("Yolanda","Ndlovu",7,True,False,True,None),
    ("Aaron","Mpundu",7,True,True,False,None),
    ("Bertha","Kafwanka",7,True,False,True,None),
    ("Victor","Mumba",6,True,True,True,None),
]

fi = 0
for term_label, due_date in terms:
    term_idx = terms.index((term_label, due_date))
    for stu in fee_students_sample:
        fname, lname, grade, has_tuition, has_transport, has_meals, special = stu
        person = f"{fname} {lname}"
        ref_base = f"FEE-{term_idx+1}-{fname[:3].upper()}{lname[:3].upper()}"
        # Tuition
        if has_tuition:
            amount = 2500
            if special == "partial":
                shortfall = 500 + (term_idx * 200)
                paid = amount - shortfall
                status = "partial"
                note = f"HS-2: Amara Dube Term {term_idx+1} partial — shortfall ZMW {shortfall}"
                fee_highlights[fi] = ERR
            elif special == "double" and term_idx == 1:
                paid = amount
                status = "paid"
                note = "HS-12: Duplicate payment — fee paid twice this term (clerical error)"
                fee_rows.append(("income","tuition",ref_base+"-T-DUP",amount,amount,"paid",due_date,person,f"Tuition {term_label} — DUPLICATE",note))
                fee_highlights[fi+1] = WARN
                fi += 1
                note = "HS-12: See duplicate above"
                fee_highlights[fi] = WARN
            else:
                paid = amount; status = "paid"; note = ""
            fee_rows.append(("income","tuition",ref_base+"-T",amount,paid,status,due_date,person,f"Tuition {term_label}",note))
            fi += 1
        # Transport
        if has_transport:
            fee_rows.append(("income","transport",ref_base+"-TR",800,800,"paid",due_date,person,f"Transport {term_label}",""))
            fi += 1
        # Meals
        if has_meals:
            fee_rows.append(("income","meals",ref_base+"-M",600,600,"paid",due_date,person,f"Meals {term_label}",""))
            fi += 1

# Overdue: a few students with unpaid Term 2 fees
for fn, ln in [("Martin","Banda"),("Laura","Zulu"),("Oscar","Phiri")]:
    fee_rows.append(("income","tuition",f"FEE-2-{fn[:3].upper()}{ln[:3].upper()}-T",2500,0,"unpaid","2026-04-30",f"{fn} {ln}","Tuition Term 2 2026","Overdue — no payment received"))
    fee_highlights[fi] = ERR
    fi += 1

make_sheet("Transactions_Fees", FEE_HEADERS, fee_rows, highlight_rows=fee_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9 — TRANSACTIONS: PAYROLL (staff salaries Jan-Mar 2026)
# ═══════════════════════════════════════════════════════════════════════════
PAY_HEADERS = [
    "transaction_type","category","reference_number","amount","amount_paid",
    "payment_status","due_date","person_name","description","notes"
]
payroll_rows = []
pay_highlights = {}
# salary map (name: monthly amount)
salary_map = {
    "Margaret Banda":15000, "Charles Mwansa":12000, "Grace Phiri":9000,
    "Joseph Ndlovu":9000, "Esther Tembo":9000, "Daniel Zulu":9000,
    "Patricia Chirwa":9000, "Simon Tembo":9000, "Loveness Mulenga":9000,
    "Francis Kabwe":8500, "Agnes Mwale":8500, "Kelvin Lungu":8500,
    "Beatrice Sakala":10000, "Peter Chanda":7000, "Naomi Mwamba":7500,
    "Moses Bwalya":8000, "Rose Nkonde":8500, "James Mpundu":4000,
    "Alice Musonda":4500, "Henry Kafwanka":4500, "Edwin Mwila":6000,
    "Godfrey Siame":6000, "Miriam Banda":4500, "Florence Mutale":4000,
    "Bertha Ngoma":4000, "Elias Mumba":3500, "Susan Kaputo":7500,
    "Charity Mbewe":3500, "Victor Chisanga":9500,
}
months = [("Jan 2026","2026-01-31"),("Feb 2026","2026-02-28"),("Mar 2026","2026-03-31")]
pi = 0
for m_label, m_due in months:
    for name, salary in salary_map.items():
        ref = f"SAL-{m_label.replace(' ','-')}-{name.replace(' ','-')[:10]}"
        note = ""
        flag = None
        if name == "Joseph Ndlovu" and m_label == "Feb 2026":
            payroll_rows.append(("expense","salary",ref+"-DUP",salary,salary,"paid",m_due,name,f"Salary {m_label} — DUPLICATE ENTRY","HS-3: Joseph Ndlovu paid TWICE in February 2026 — reconciliation required"))
            pay_highlights[pi] = ERR
            pi += 1
            note = "HS-3: Original entry — also see duplicate above"
            flag = WARN
        payroll_rows.append(("expense","salary",ref,salary,salary,"paid",m_due,name,f"Salary {m_label}",note))
        if flag:
            pay_highlights[pi] = flag
        pi += 1

make_sheet("Transactions_Payroll", PAY_HEADERS, payroll_rows, highlight_rows=pay_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 10 — TRANSACTIONS: PROCUREMENT
# ═══════════════════════════════════════════════════════════════════════════
PROC_HEADERS = [
    "transaction_type","category","reference_number","amount","amount_paid",
    "payment_status","due_date","person_name","description","notes"
]
procurement_rows = []
proc_highlights = {}
procurement = [
    # EduSupply (HS-6: paid but no goods receipt)
    ("expense","purchase_order","PO-2026-001",85000,85000,"paid","2026-01-20","David Mutale","Textbooks — EduSupply Zambia","HS-6: Paid 2026-01-20. Goods receipt task still INCOMPLETE",ERR),
    ("expense","purchase_order","PO-2026-002",18500,18500,"paid","2026-01-25","David Mutale","Art Supplies — EduSupply Zambia","HS-6: Paid 2026-01-25. Goods receipt task INCOMPLETE — stock=0",ERR),
    ("expense","purchase_order","PO-2026-003",12000,12000,"paid","2026-02-01","David Mutale","Lab Supplies — EduSupply Zambia","HS-6: Paid 2026-02-01. Expired Ethanol accepted in goods receipt",WARN),
    # Normal suppliers
    ("expense","purchase_order","PO-2026-004",8500,8500,"paid","2026-01-15","Christine Bwalya","Stationery Q1 — Office Plus","",None),
    ("expense","purchase_order","PO-2026-005",6200,6200,"paid","2026-01-18","Christine Bwalya","Exercise books and pens — Office Plus","",None),
    ("expense","purchase_order","PO-2026-006",24500,24500,"paid","2026-02-05","Robert Phiri","Computer lab peripherals — Tech Supplies","Mouse and keyboards for computer lab",None),
    ("expense","purchase_order","PO-2026-007",1200,1200,"paid","2026-02-12","Nancy Sakala","First aid restock — MedSupply","Bandages, antiseptic",None),
    ("expense","purchase_order","PO-2026-008",3500,0,"unpaid","2026-03-31","Christine Bwalya","Stationery Q2 — Office Plus","Awaiting delivery confirmation",WARN),
    # Maintenance
    ("expense","maintenance","MNT-2026-001",4500,4500,"paid","2026-01-20","","Bus 2 service — Toyota dealer","",None),
    ("expense","maintenance","MNT-2026-002",0,0,"unpaid","2026-01-31","","Bus 3 overdue service — NOT YET BOOKED","HS-1: Service overdue — 0 spent, not yet arranged",ERR),
    ("expense","maintenance","MNT-2026-003",2200,2200,"paid","2026-02-15","","Projector lamp purchase attempt","Ordered wrong part — lamp still needed",WARN),
    # Utilities
    ("expense","utilities","UTIL-2026-001",3200,3200,"paid","2026-01-31","","Electricity — January 2026","",None),
    ("expense","utilities","UTIL-2026-002",3400,3400,"paid","2026-02-28","","Electricity — February 2026","",None),
    ("expense","utilities","UTIL-2026-003",3100,3100,"paid","2026-03-31","","Electricity — March 2026","",None),
    ("expense","utilities","WATER-2026-001",1500,1500,"paid","2026-01-31","","Water — January 2026","",None),
    ("expense","utilities","WATER-2026-002",1600,1600,"paid","2026-02-28","","Water — February 2026","",None),
    # Security contract
    ("expense","services","SVC-2026-001",8000,8000,"paid","2026-01-31","Arnold Mwamba","Security services — Jan 2026","",None),
    ("expense","services","SVC-2026-002",8000,8000,"paid","2026-02-28","Arnold Mwamba","Security services — Feb 2026","",None),
    ("expense","services","SVC-2026-003",8000,8000,"paid","2026-03-31","Arnold Mwamba","Security services — Mar 2026","",None),
]
for ri, rec in enumerate(procurement_rows.__class__()):
    pass
for ri, rec in enumerate(procurement):
    flag = rec[-1]
    row = rec[:-1]
    procurement_rows.append(row)
    if flag:
        proc_highlights[ri] = flag

make_sheet("Transactions_Procurement", PROC_HEADERS, procurement_rows, highlight_rows=proc_highlights)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 11 — RELATIONSHIPS
# ═══════════════════════════════════════════════════════════════════════════
REL_HEADERS = [
    "entity_a_type","entity_a_name","relationship_type","entity_b_type","entity_b_name","notes"
]
relationships = [
    # Task dependencies (HS-11 exam sequence)
    ("task","Grade 7 Exam Invigilation — Session 1","depends_on","task","Grade 7 Exam Paper Printing","HS-11: Invigilation cannot run before papers are printed"),
    ("task","Grade 7 Exam Invigilation — Session 2","depends_on","task","Grade 7 Exam Invigilation — Session 1","Session 2 after Session 1"),
    ("task","Grade 7 Exam Results Compilation","depends_on","task","Grade 7 Exam Invigilation — Session 2","Results after all invigilation complete"),
    ("task","Grade 7 Exam Paper Printing","depends_on","task","Grade 7 Exam Schedule Confirmed","Papers can only be printed after schedule is confirmed"),
    # Procurement goods receipt dependencies (HS-6)
    ("task","Goods Receipt — EduSupply Textbooks","depends_on","task","Procurement Order — EduSupply (Textbooks)","HS-6: Goods receipt should follow purchase order"),
    ("task","Goods Receipt — EduSupply Art Supplies","depends_on","task","Procurement Order — EduSupply (Art Supplies)","HS-6: Goods receipt should follow purchase order"),
    # Field trip dependency (HS-7)
    ("task","Grade 6 Field Trip — Science Museum Lusaka","depends_on","task","Grade 6 Field Trip Parent Consent Collection","All consents required before trip proceeds"),
    ("task","Grade 6 Field Trip — Science Museum Lusaka","depends_on","task","Grade 6 Field Trip Risk Assessment","Risk assessment required before trip"),
    # Bus 3 dependency
    ("task","Bus 3 (AAM 1198) — Full Service Overdue","blocks","task","Grade 6 Field Trip — Science Museum Lusaka","HS-1: Bus 3 should NOT be used for field trip until serviced"),
    # Staff — class assignments
    ("person","Grace Phiri","teaches","enterprise","Grade 1-3 Block","Grade 1 class teacher"),
    ("person","Joseph Ndlovu","teaches","enterprise","Grade 1-3 Block","Grade 2 class teacher"),
    ("person","Esther Tembo","teaches","enterprise","Grade 1-3 Block","Grade 3 class teacher"),
    ("person","Daniel Zulu","teaches","enterprise","Grade 4-6 Block","Grade 4 class teacher"),
    ("person","Patricia Chirwa","teaches","enterprise","Grade 4-6 Block","Grade 5 class teacher — currently on leave"),
    ("person","Simon Tembo","teaches","enterprise","Grade 4-6 Block","Grade 6 class teacher"),
    ("person","Loveness Mulenga","teaches","enterprise","Grade 7 Block","Grade 7 class teacher"),
    ("person","Miriam Banda","covers","person","Patricia Chirwa","Relief teacher covering Grade 5 absence — HS-7 skills gap"),
    # Supervision
    ("person","Margaret Banda","supervises","person","Charles Mwansa","Principal → Vice Principal"),
    ("person","Charles Mwansa","supervises","person","Grace Phiri","Deputy → Junior Section Lead"),
    ("person","Charles Mwansa","supervises","person","Loveness Mulenga","Deputy → Senior Section Lead"),
    ("person","Beatrice Sakala","supervises","person","Peter Chanda","Bursar → Admin Clerk"),
    # Student — guardian (sample)
    ("person","Amara Dube","guardian_is","person","Alice Tembo","HS-2: PTA Chair is guardian of partial fee payer"),
    ("person","Chanda Mwale","guardian_is","person","Brenda Mulenga","HS-12: Grade 7 parent rep is guardian of double fee student"),
    # Staff — skills (for skills-based routing)
    ("person","Francis Kabwe","certified_in","task","Grade 6 Field Trip — Science Museum Lusaka","HS-7: Kabwe has first_aid — qualified lead for field trip"),
    ("person","Miriam Banda","not_certified_for","task","Grade 6 Field Trip — Science Museum Lusaka","HS-7: No first_aid — should not be sole adult on field trip"),
    ("person","Rose Nkonde","certified_in","task","Staff CPR Refresher Training","Trainer — Registered Nurse"),
    # Bus assignments
    ("person","Edwin Mwila","operates","product","School Bus 3 — Reg AAM 1198","HS-1: Senior driver assigned to most problematic bus"),
    ("person","Edwin Mwila","operates","product","School Bus 1 — Reg BAA 4521","Primary route driver"),
    ("person","Godfrey Siame","operates","product","School Bus 2 — Reg BAB 7733","HS Note: PDP expired"),
    # Product — enterprise location
    ("product","School Bus 3 — Reg AAM 1198","belongs_to","enterprise","Transport Department","HS-1: Overdue service bus"),
    ("product","Grade 7 Mathematics Textbook","belongs_to","enterprise","Grade 7 Block","HS-4: Critical low stock"),
    ("product","Ethanol 95% (1L bottle)","belongs_to","enterprise","Science & Computer Lab","HS-5: Expired chemical"),
    # Donor relations
    ("person","Mary Banda","funds","enterprise","Greenfield Primary School","UNICEF supplies donation"),
    ("person","Thomas Mwale","funds","enterprise","Grade 1-3 Block","Save the Children reading program"),
    # Supplier — enterprise
    ("person","David Mutale","supplies","enterprise","Greenfield Primary School","HS-6: EduSupply — unreliable delivery"),
    # MoE oversight
    ("person","Jennifer Kabwe","regulates","enterprise","Greenfield Primary School","District Education Officer — annual inspection"),
    # Attendance correlations (for copilot query)
    ("person","Patricia Chirwa","impacts","enterprise","Grade 4-6 Block","HS-10: Chirwa absences correlate with Grade 5 outcome drop"),
]
make_sheet("Relationships", REL_HEADERS, relationships, highlight_rows={0:ERR,4:ERR,5:ERR,8:ERR,15:WARN,16:WARN,22:WARN,25:WARN,27:WARN,28:ERR,35:WARN})

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 12 — ADDRESSES
# ═══════════════════════════════════════════════════════════════════════════
ADDR_HEADERS = [
    "address_type","label","street","city","region","country","postal_code","notes"
]
addresses = [
    ("operational","Greenfield Primary School — Main Campus","Plot 14 Leopards Hill Road","Lusaka","Lusaka Province","Zambia","10101","Main school site — 3 hectares"),
    ("operational","Grade 1-3 Block","Plot 14 Leopards Hill Road — Block A","Lusaka","Lusaka Province","Zambia","10101","Junior section — classrooms 1-3"),
    ("operational","Grade 4-6 Block","Plot 14 Leopards Hill Road — Block B","Lusaka","Lusaka Province","Zambia","10101","Middle section — classrooms 4-6"),
    ("operational","Grade 7 Block","Plot 14 Leopards Hill Road — Block C","Lusaka","Lusaka Province","Zambia","10101","Senior section — classroom 7 + exam hall"),
    ("operational","Administration Block","Plot 14 Leopards Hill Road — Admin Wing","Lusaka","Lusaka Province","Zambia","10101","Principal office, bursar, staffroom"),
    ("operational","Science & Computer Lab","Plot 14 Leopards Hill Road — Lab Block","Lusaka","Lusaka Province","Zambia","10101","HS-5: Expired chemicals stored here"),
    ("operational","Transport Yard","Plot 14 Leopards Hill Road — Transport Bay","Lusaka","Lusaka Province","Zambia","10101","HS-1: Bus 3 parked here awaiting service"),
    ("supplier","EduSupply Zambia Ltd — Warehouse","45 Great East Road Industrial","Lusaka","Lusaka Province","Zambia","10102","HS-6: Supplier with unreliable delivery"),
    ("supplier","Office Plus Zambia","23 Cairo Road","Lusaka","Lusaka Province","Zambia","10103","Stationery supplier"),
    ("supplier","Tech Supplies Africa","89 Freedom Way","Lusaka","Lusaka Province","Zambia","10104","Computer supplier"),
    ("government","Ministry of Education — Lusaka District Office","Government Complex, Ridgeway","Lusaka","Lusaka Province","Zambia","10001","Annual school reporting"),
    ("government","Road Transport Safety Authority","RTSA House, Cairo Road","Lusaka","Lusaka Province","Zambia","10002","Bus inspection authority"),
    ("residential","Edwin Mwila — Home Address","House 3, Chalala Township","Lusaka","Lusaka Province","Zambia","10201","Senior bus driver — HS-1 related"),
    ("other","Science Museum Lusaka","Freedom Way, Longacres","Lusaka","Lusaka Province","Zambia","10105","HS-7: Grade 6 field trip destination"),
]
make_sheet("Addresses", ADDR_HEADERS, addresses, highlight_rows={5:WARN,6:WARN,7:ERR})

# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════
outpath = r"c:\Users\write\OneDrive\Desktop\Newsconseen\2. Newsconseen Developer\newsconseenwebapp\Greenfield_Primary_School_Demo_Data.xlsx"
WB.save(outpath)
print(f"Saved: {outpath}")
print(f"Sheets: {[s.title for s in WB.worksheets]}")
total = sum(WB[s.title].max_row - 1 for s in WB.worksheets if WB[s.title].max_row > 1)
print(f"Approx data rows: {total}")
