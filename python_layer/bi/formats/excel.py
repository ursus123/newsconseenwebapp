"""
bi/formats/excel.py
--------------------
Generates a multi-sheet .xlsx file from a report dict.

Power BI imports this in one click: Get Data → Excel → select file.
Excel opens it directly. Google Sheets imports it for Looker Studio.

Formatting:
  - Bold, coloured header row (Newsconseen brand blue)
  - Frozen first row so headers stay visible while scrolling
  - Auto-filter on every column
  - Column widths auto-sized (capped at 40 chars)
  - "Meta" sheet appended last with export metadata
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import pandas as pd


# Brand colours
_HEADER_BG   = "1E40AF"   # Newsconseen blue-800
_HEADER_FONT = "FFFFFF"   # white text
_ALT_ROW_BG  = "EFF6FF"   # blue-50 for alternating rows


def _write_sheet(ws, df: pd.DataFrame, sheet_title: str) -> None:
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    if df.empty:
        ws.append(["No data available for this report."])
        return

    cols = list(df.columns)

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(bold=True, color=_HEADER_FONT, size=10)

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=_ALT_ROW_BG)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            # Convert None/NaN to empty string; booleans to Yes/No
            if value is None or (isinstance(value, float) and value != value):
                value = ""
            elif isinstance(value, bool):
                value = "Yes" if value else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(cols, start=1):
        letter = get_column_letter(col_idx)
        col_data = df.iloc[:, col_idx - 1].astype(str)
        max_len  = max(len(str(col_name)), col_data.str.len().max() if not col_data.empty else 0)
        ws.column_dimensions[letter].width = min(max_len + 3, 42)

    # ── Freeze top row + auto-filter ──────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _meta_sheet(wb: openpyxl.Workbook, report_title: str, description: str, sheets: list[dict]) -> None:
    """Append a 'Meta' sheet with export information."""
    ws = wb.create_sheet("Meta")
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(bold=True, color=_HEADER_FONT)

    rows: list[tuple[str, Any]] = [
        ("Report",       report_title),
        ("Description",  description),
        ("Exported at",  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Source",       "Newsconseen Analytics Layer (analytics.*)"),
        ("Power BI tip", "Home → Get Data → Excel → select this file"),
        ("Looker tip",   "Import to Google Sheets → connect as Looker Studio data source"),
        ("Sheets",       ", ".join(s["name"] for s in sheets if not s["df"].empty)),
    ]

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 65

    for r_idx, (key, val) in enumerate(rows, start=1):
        ka = ws.cell(row=r_idx, column=1, value=key)
        ka.fill = header_fill
        ka.font = header_font
        ws.cell(row=r_idx, column=2, value=str(val))


def build_excel(report: dict) -> bytes:
    """
    Convert a report dict → .xlsx bytes.

    Parameters
    ----------
    report  dict from generators.py — keys: title, description, sheets, primary
    """
    wb = openpyxl.Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    for sheet_spec in report["sheets"]:
        ws = wb.create_sheet(title=sheet_spec["name"][:31])  # Excel tab name limit = 31 chars
        _write_sheet(ws, sheet_spec["df"], sheet_spec["name"])

    _meta_sheet(wb, report["title"], report["description"], report["sheets"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
