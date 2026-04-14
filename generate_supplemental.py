"""
Greenfield Primary School — Supplemental Import Workbook
4 sheets mapped exactly to Newsconseen's importer field names:
  1. Relationships_Import  — remapped from entity_a/b to person/enterprise/item_name
  2. Transactions_Fees     — school fee income with date column added
  3. Transactions_Payroll  — staff salaries with date column added
  4. Transactions_Procurement — expenses/procurement with date column added
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

WB = openpyxl.Workbook()
WB.remove(WB.active)

# ── Colour palette ─────────────────────────────────────────────────────────
HDR  = PatternFill("solid", fgColor="1E293B")   # slate-900 header
ERR  = PatternFill("solid", fgColor="FEE2E2")   # red-100
WARN = PatternFill("solid", fgColor="FEF9C3")   # yellow-100
SKIP = PatternFill("solid", fgColor="F1F5F9")   # slate-100 skipped rows

HDR_FONT  = Font(color="FFFFFF", bold=True)
SKIP_FONT = Font(color="94A3B8", italic=True)

def make_sheet(title, headers, rows, highlight_rows=None, skip_rows=None):
    ws = WB.create_sheet(title=title)
    highlight_rows = highlight_rows or {}
    skip_rows = skip_rows or set()

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HDR
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(h) + 4)

    # Data rows
    for ri, row in enumerate(rows):
        excel_row = ri + 2
        is_skip = ri in skip_rows
        fill = highlight_rows.get(ri, None)
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            if is_skip:
                c.fill = SKIP
                c.font = SKIP_FONT
            elif fill:
                c.fill = fill

    ws.freeze_panes = "A2"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — RELATIONSHIPS_IMPORT
# Mapped to Newsconseen fields: relationship_type, person_name,
# enterprise_name, item_name, role, start_date, status, notes
#
# Skipped (no supported type): task→task, person→task
# person→person uses person_name (A) + enterprise_name (B) — best available
# ═══════════════════════════════════════════════════════════════════════════
REL_HEADERS = [
    "relationship_type",
    "person_name",
    "enterprise_name",
    "item_name",
    "role",
    "start_date",
    "status",
    "notes",
]

# fmt: off
# (relationship_type, person_name, enterprise_name, item_name, role, start_date, status, notes)
relationships = [
    # ── Staff class assignments (person_enterprise) ─────────────────────
    ("person_enterprise","Grace Phiri",       "Grade 1-3 Block",          "","teaches",    "2024-01-15","active","Grade 1 class teacher"),
    ("person_enterprise","Joseph Ndlovu",     "Grade 1-3 Block",          "","teaches",    "2024-01-15","active","Grade 2 class teacher"),
    ("person_enterprise","Esther Tembo",      "Grade 1-3 Block",          "","teaches",    "2024-01-15","active","Grade 3 class teacher"),
    ("person_enterprise","Daniel Zulu",       "Grade 4-6 Block",          "","teaches",    "2024-01-15","active","Grade 4 class teacher"),
    ("person_enterprise","Patricia Chirwa",   "Grade 4-6 Block",          "","teaches",    "2024-01-15","active","Grade 5 class teacher — currently on leave"),
    ("person_enterprise","Simon Tembo",       "Grade 4-6 Block",          "","teaches",    "2024-01-15","active","Grade 6 class teacher"),
    ("person_enterprise","Loveness Mulenga",  "Grade 7 Block",            "","teaches",    "2024-01-15","active","Grade 7 class teacher"),
    # ── Person-to-person: coverage & supervision (person_person) ────────
    # Note: enterprise_name field used for the second person (B side)
    ("person_person",    "Miriam Banda",      "Patricia Chirwa",          "","covers",     "2026-01-20","active","HS-7: Relief teacher covering Grade 5 absence — skills gap"),
    ("person_person",    "Margaret Banda",    "Charles Mwansa",           "","supervises", "2024-01-15","active","Principal supervises Vice Principal"),
    ("person_person",    "Charles Mwansa",    "Grace Phiri",              "","supervises", "2024-01-15","active","Deputy supervises Junior Section Lead"),
    ("person_person",    "Charles Mwansa",    "Loveness Mulenga",         "","supervises", "2024-01-15","active","Deputy supervises Senior Section Lead"),
    ("person_person",    "Beatrice Sakala",   "Peter Chanda",             "","supervises", "2024-01-15","active","Bursar supervises Admin Clerk"),
    ("person_person",    "Amara Dube",        "Alice Tembo",              "","guardian_is","2024-01-15","active","HS-2: PTA Chair is guardian of partial fee payer"),
    ("person_person",    "Chanda Mwale",      "Brenda Mulenga",           "","guardian_is","2024-01-15","active","HS-12: Grade 7 parent rep is guardian of double fee student"),
    # ── Bus driver assignments (person_item) ────────────────────────────
    ("person_item",      "Edwin Mwila",       "","School Bus 3 — Reg AAM 1198","operates","2023-06-01","active","HS-1: Senior driver assigned to most problematic bus"),
    ("person_item",      "Edwin Mwila",       "","School Bus 1 — Reg BAA 4521","operates","2023-06-01","active","Primary route driver"),
    ("person_item",      "Godfrey Siame",     "","School Bus 2 — Reg BAB 7733","operates","2024-01-15","active","HS Note: PDP expired"),
    # ── Product-enterprise location (item_enterprise) ───────────────────
    ("item_enterprise",  "","Transport Department",     "School Bus 3 — Reg AAM 1198",   "belongs_to","2023-06-01","active","HS-1: Overdue service bus"),
    ("item_enterprise",  "","Grade 7 Block",            "Grade 7 Mathematics Textbook",  "belongs_to","2024-01-15","active","HS-4: Critical low stock"),
    ("item_enterprise",  "","Science & Computer Lab",   "Ethanol 95% (1L bottle)",       "belongs_to","2024-01-15","active","HS-5: Expired chemical"),
    # ── Donor and external relations (person_enterprise) ────────────────
    ("person_enterprise","Mary Banda",        "Greenfield Primary School", "","funds",     "2025-01-10","active","UNICEF supplies donation"),
    ("person_enterprise","Thomas Mwale",      "Grade 1-3 Block",          "","funds",     "2025-03-01","active","Save the Children reading program"),
    ("person_enterprise","David Mutale",      "Greenfield Primary School", "","supplies",  "2025-01-01","active","HS-6: EduSupply — unreliable delivery"),
    ("person_enterprise","Jennifer Kabwe",    "Greenfield Primary School", "","regulates", "2024-06-01","active","District Education Officer — annual inspection"),
    ("person_enterprise","Patricia Chirwa",   "Grade 4-6 Block",          "","impacts",   "2026-01-20","active","HS-10: Chirwa absences correlate with Grade 5 outcome drop"),
]
# fmt: on

rel_highlights = {
    3:  WARN,   # Patricia Chirwa on leave
    7:  WARN,   # Miriam covering (skills gap)
    12: WARN,   # PTA chair guardian of partial payer
    13: WARN,   # parent rep guardian of double payment
    14: ERR,    # Bus 3 overdue service
    17: ERR,    # Bus 3 belongs_to Transport Dept (highlight risk)
    18: WARN,   # Low stock textbook
    19: ERR,    # Expired chemical
    22: ERR,    # EduSupply unreliable
    23: WARN,   # Chirwa impacts Grade 4-6
}

make_sheet("Relationships_Import", REL_HEADERS, relationships, highlight_rows=rel_highlights)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — TRANSACTIONS_FEES
# Added: date column (= due_date — when fee was invoiced/due)
# Added: enterprise column (Greenfield Primary School)
# Mapped: person_name → primary_person
# status = posted (historical records)
# ═══════════════════════════════════════════════════════════════════════════
FEE_HEADERS = [
    "transaction_type","date","reference_number","amount","amount_paid",
    "payment_status","status","due_date","enterprise","primary_person",
    "description","notes"
]

# term → (label, date, due_date)
TERMS = [
    ("Term 1 2026", "2026-01-15", "2026-01-31"),
    ("Term 2 2026", "2026-04-07", "2026-04-30"),
    ("Term 3 2026", "2026-08-01", "2026-08-31"),
]

FEE_STUDENTS = [
    # (first, last, grade, has_tuition, has_transport, has_meals, special)
    ("Amara",   "Dube",      7, True, True,  True,  "partial"),
    ("Chanda",  "Mwale",     7, True, False, True,  "double"),
    ("Winnie",  "Chirwa",    7, True, True,  True,  None),
    ("Timothy", "Chanda",    7, True, True,  True,  None),
    ("Yolanda", "Ndlovu",    7, True, False, True,  None),
    ("Aaron",   "Mpundu",    7, True, True,  False, None),
    ("Bertha",  "Kafwanka",  7, True, False, True,  None),
    ("Victor",  "Mumba",     6, True, True,  True,  None),
    ("Martin",  "Phiri",     5, True, True,  False, None),
    ("Laura",   "Zulu",      4, True, False, True,  None),
]

fee_rows = []
fee_highlights = {}
fi = 0
ENTERPRISE = "Greenfield Primary School"

for term_label, tx_date, due_date in TERMS:
    term_idx = TERMS.index((term_label, tx_date, due_date))
    for stu in FEE_STUDENTS:
        fname, lname, grade, has_tuition, has_transport, has_meals, special = stu
        person = f"{fname} {lname}"
        ref_base = f"FEE-{term_idx+1}-{fname[:3].upper()}{lname[:3].upper()}"

        if has_tuition:
            amount = 2500
            if special == "partial":
                shortfall = 500 + (term_idx * 200)
                paid = amount - shortfall
                pstatus = "partial"
                note = f"HS-2: {person} Term {term_idx+1} partial — shortfall ZMW {shortfall}"
                fee_highlights[fi] = ERR
            elif special == "double" and term_idx == 1:
                # Duplicate entry
                fee_rows.append((
                    "sale_service", tx_date, ref_base+"-T-DUP", amount, amount,
                    "paid", "posted", due_date, ENTERPRISE, person,
                    f"Tuition {term_label} — DUPLICATE", "HS-12: Duplicate payment — fee paid twice this term"
                ))
                fee_highlights[fi] = ERR
                fi += 1
                paid = amount; pstatus = "paid"
                note = "HS-12: See duplicate entry above"
                fee_highlights[fi] = WARN
            else:
                paid = amount; pstatus = "paid"; note = ""
            fee_rows.append((
                "sale_service", tx_date, ref_base+"-T", amount, paid,
                pstatus, "posted", due_date, ENTERPRISE, person,
                f"Tuition {term_label}", note
            ))
            fi += 1

        if has_transport:
            fee_rows.append((
                "sale_service", tx_date, ref_base+"-TR", 800, 800,
                "paid", "posted", due_date, ENTERPRISE, person,
                f"Transport {term_label}", ""
            ))
            fi += 1

        if has_meals:
            fee_rows.append((
                "sale_service", tx_date, ref_base+"-M", 600, 600,
                "paid", "posted", due_date, ENTERPRISE, person,
                f"Meals {term_label}", ""
            ))
            fi += 1

# Overdue Term 2 fees
for fn, ln in [("Martin","Banda"),("Laura","Zulu"),("Oscar","Phiri")]:
    fee_rows.append((
        "sale_service", "2026-04-07",
        f"FEE-2-{fn[:3].upper()}{ln[:3].upper()}-T",
        2500, 0, "unpaid", "posted", "2026-04-30",
        ENTERPRISE, f"{fn} {ln}", "Tuition Term 2 2026",
        "Overdue — no payment received"
    ))
    fee_highlights[fi] = ERR
    fi += 1

make_sheet("Transactions_Fees", FEE_HEADERS, fee_rows, highlight_rows=fee_highlights)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — TRANSACTIONS_PAYROLL
# ═══════════════════════════════════════════════════════════════════════════
PAY_HEADERS = [
    "transaction_type","date","reference_number","amount","amount_paid",
    "payment_status","status","due_date","enterprise","primary_person",
    "description","notes"
]

SALARY_MAP = {
    "Margaret Banda":15000, "Charles Mwansa":12000, "Grace Phiri":9000,
    "Joseph Ndlovu":9000,   "Esther Tembo":9000,    "Daniel Zulu":9000,
    "Patricia Chirwa":9000, "Simon Tembo":9000,     "Loveness Mulenga":9000,
    "Francis Kabwe":8500,   "Agnes Mwale":8500,     "Kelvin Lungu":8500,
    "Beatrice Sakala":10000,"Peter Chanda":7000,    "Naomi Mwamba":7500,
    "Moses Bwalya":8000,    "Rose Nkonde":8500,     "James Mpundu":4000,
    "Alice Musonda":4500,   "Henry Kafwanka":4500,  "Edwin Mwila":6000,
    "Godfrey Siame":6000,   "Miriam Banda":4500,    "Florence Mutale":4000,
    "Bertha Ngoma":4000,    "Elias Mumba":3500,     "Susan Kaputo":7500,
    "Charity Mbewe":3500,   "Victor Chisanga":9500,
}

MONTHS = [
    ("Jan 2026","2026-01-31","2026-01-31"),
    ("Feb 2026","2026-02-28","2026-02-28"),
    ("Mar 2026","2026-03-31","2026-03-31"),
]

payroll_rows = []
pay_highlights = {}
pi = 0

for m_label, m_date, m_due in MONTHS:
    for name, salary in SALARY_MAP.items():
        ref = f"SAL-{m_label.replace(' ','-')}-{name.replace(' ','-')[:10]}"
        if name == "Joseph Ndlovu" and m_label == "Feb 2026":
            # Duplicate salary entry
            payroll_rows.append((
                "expense", m_date, ref+"-DUP", salary, salary,
                "paid", "posted", m_due, ENTERPRISE, name,
                f"Salary {m_label} — DUPLICATE ENTRY",
                "HS-3: Joseph Ndlovu paid TWICE in February 2026 — reconciliation required"
            ))
            pay_highlights[pi] = ERR
            pi += 1
            note = "HS-3: Original entry — also see duplicate above"
            pay_highlights[pi] = WARN
        else:
            note = ""
        payroll_rows.append((
            "expense", m_date, ref, salary, salary,
            "paid", "posted", m_due, ENTERPRISE, name,
            f"Salary {m_label}", note
        ))
        pi += 1

make_sheet("Transactions_Payroll", PAY_HEADERS, payroll_rows, highlight_rows=pay_highlights)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — TRANSACTIONS_PROCUREMENT
# ═══════════════════════════════════════════════════════════════════════════
PROC_HEADERS = [
    "transaction_type","date","reference_number","amount","amount_paid",
    "payment_status","status","due_date","enterprise","primary_person",
    "description","notes"
]

# (tx_type, date, ref, amount, amount_paid, pstatus, due_date, person, desc, notes, highlight)
procurement_raw = [
    ("expense","2026-01-20","PO-2026-001",85000,85000,"paid","2026-01-20","David Mutale","Textbooks — EduSupply Zambia","HS-6: Paid 2026-01-20. Goods receipt task still INCOMPLETE",ERR),
    ("expense","2026-01-25","PO-2026-002",18500,18500,"paid","2026-01-25","David Mutale","Art Supplies — EduSupply Zambia","HS-6: Paid 2026-01-25. Goods receipt task INCOMPLETE — stock=0",ERR),
    ("expense","2026-02-01","PO-2026-003",12000,12000,"paid","2026-02-01","David Mutale","Lab Supplies — EduSupply Zambia","HS-6: Paid 2026-02-01. Expired Ethanol accepted in goods receipt",WARN),
    ("expense","2026-01-15","PO-2026-004",8500, 8500, "paid","2026-01-15","Christine Bwalya","Stationery Q1 — Office Plus","",None),
    ("expense","2026-01-18","PO-2026-005",6200, 6200, "paid","2026-01-18","Christine Bwalya","Exercise books and pens — Office Plus","",None),
    ("expense","2026-02-05","PO-2026-006",24500,24500,"paid","2026-02-05","Robert Phiri","Computer lab peripherals — Tech Supplies","Mouse and keyboards for computer lab",None),
    ("expense","2026-02-12","PO-2026-007",1200, 1200, "paid","2026-02-12","Nancy Sakala","First aid restock — MedSupply","Bandages, antiseptic",None),
    ("expense","2026-03-31","PO-2026-008",3500, 0,    "unpaid","2026-03-31","Christine Bwalya","Stationery Q2 — Office Plus","Awaiting delivery confirmation",WARN),
    ("expense","2026-01-20","MNT-2026-001",4500,4500, "paid","2026-01-20","","Bus 2 service — Toyota dealer","",None),
    ("expense","2026-01-31","MNT-2026-002",0,   0,    "unpaid","2026-01-31","","Bus 3 overdue service — NOT YET BOOKED","HS-1: Service overdue — 0 spent, not yet arranged",ERR),
    ("expense","2026-02-15","MNT-2026-003",2200,2200, "paid","2026-02-15","","Projector lamp purchase attempt","Ordered wrong part — lamp still needed",WARN),
    ("expense","2026-01-31","UTIL-2026-001",3200,3200,"paid","2026-01-31","","Electricity — January 2026","",None),
    ("expense","2026-02-28","UTIL-2026-002",3400,3400,"paid","2026-02-28","","Electricity — February 2026","",None),
    ("expense","2026-03-31","UTIL-2026-003",3100,3100,"paid","2026-03-31","","Electricity — March 2026","",None),
    ("expense","2026-01-31","WATER-2026-001",1500,1500,"paid","2026-01-31","","Water — January 2026","",None),
    ("expense","2026-02-28","WATER-2026-002",1600,1600,"paid","2026-02-28","","Water — February 2026","",None),
    ("expense","2026-01-31","SVC-2026-001",8000,8000, "paid","2026-01-31","Arnold Mwamba","Security services — Jan 2026","",None),
    ("expense","2026-02-28","SVC-2026-002",8000,8000, "paid","2026-02-28","Arnold Mwamba","Security services — Feb 2026","",None),
    ("expense","2026-03-31","SVC-2026-003",8000,8000, "paid","2026-03-31","Arnold Mwamba","Security services — Mar 2026","",None),
]

proc_rows = []
proc_highlights = {}
for ri, rec in enumerate(procurement_raw):
    tx_type,date,ref,amt,paid,pstatus,due,person,desc,notes,flag = rec
    proc_rows.append((tx_type, date, ref, amt, paid, pstatus, "posted", due, ENTERPRISE, person, desc, notes))
    if flag:
        proc_highlights[ri] = flag

make_sheet("Transactions_Procurement", PROC_HEADERS, proc_rows, highlight_rows=proc_highlights)


# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════
outpath = r"c:\Users\write\OneDrive\Desktop\Newsconseen\2. Newsconseen Developer\newsconseenwebapp\Greenfield_Supplemental_Import.xlsx"
WB.save(outpath)
print(f"Saved: {outpath}")
for ws in WB.worksheets:
    print(f"  {ws.title}: {ws.max_row - 1} data rows")
